import os
import xlsxwriter
import sys
from PIL import Image
import openpyxl

# for avoiding error while exporting Japanese
if sys.version[0] == '2':
    reload(sys)
    sys.setdefaultencoding('utf8')

# Col constant name
INDEX_NO = 0
FILE_NAME = 1
FIELD_ID = 2
LINE_IMG = 3
LINE_OCR = 4
TRUE_FALSE = 5
ACCURACY = 6

HIDDEN_COL = 100

DEFAULT_ROW_HEIGHT = 50
LINE_IMG_WIDTH = 100


def replace(parent):
    """
    :param parent:
        the path of parent dir to iterate throught
    :return:
        replace spaces ' ' and brackets '(' and ')' with underscore '_'
    """

    for path, folders, files in os.walk(parent):
        for f in files:
            os.rename(os.path.join(path, f),
                      os.path.join(path, f.replace(' ', '_').replace('(', '_').replace(')', '_')))

        for i in range(len(folders)):
            new_name = folders[i].replace(' ', '_').replace('(', '_').replace(')', '_')
            os.rename(os.path.join(path, folders[i]), os.path.join(path, new_name))
            folders[i] = new_name


def gen_report_file(prefer_name, output_path, template_path, is_single_dir=False):
    """

    :param prefer_name: prefer name of output .xls file
    :param output_path: prefer output dir of .xls file
    :param data_path: input of data
    :return:
    """
    # form_name = os.path.basename(data_path)

    # init workbook obj with preferred name
    out_filename = output_path + "/" + prefer_name + ".xlsx"
    print(out_filename)
    workbook = xlsxwriter.Workbook(out_filename)
    # add worksheets to the workbook
    worksheet1 = workbook.add_worksheet()

    # format the font in a cell
    font_size_format = workbook.add_format()
    font_size_format.set_font_size(14)
    font_size_format.set_align('center')
    font_size_format.set_align('vcenter')
    font_size_format.set_border()
    font_size_format.set_text_wrap(True)

    general_format = workbook.add_format()
    general_format.set_align('center')
    general_format.set_align('vcenter')
    general_format.set_border()
    general_format.set_text_wrap(True)

    # Create a format to use in the merged range.
    merge_format = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })

    # format for disabled text
    gray_format =  workbook.add_format({
        'font_size': 9,
        'font_color': 'gray',
        'valign': 'vcenter'
    })

    # set the width of columns in sheet 1
    worksheet1.set_column(FILE_NAME, FILE_NAME, 10)
    worksheet1.set_column(FIELD_ID, FIELD_ID, 7)
    worksheet1.set_column(LINE_IMG, LINE_IMG, LINE_IMG_WIDTH)
    worksheet1.set_column(LINE_OCR, LINE_OCR, 40)
    worksheet1.set_column(TRUE_FALSE, TRUE_FALSE, 50)

    # set default height of all row of sheet 1
    worksheet1.set_default_row(DEFAULT_ROW_HEIGHT)
    # create the title header of the sheet
    worksheet1.write(0, INDEX_NO, "#")

    worksheet1.write(0, FILE_NAME, "FILE NAME")
    worksheet1.write(0, FIELD_ID, "FIELD ID")
    worksheet1.write(0, LINE_IMG, "TEXT OR NUMBER")
    worksheet1.write(0, LINE_OCR, "LINE OCR")
    worksheet1.write(0, TRUE_FALSE, "TRUE OR FALSE")

    dashline_ext = ["line_", "dashline_number", "dashline_text"]
    checkbox_keyword = ["checkbox", "origin"]
    mixedline_keyword = "mixedline"

    # fix spaces in paths
    replace(template_path)

    global_index = 1
    if is_single_dir:
        form_dirs = [template_path]
    else:
        form_dirs = [os.path.join(template_path, name) for name in os.listdir(template_path) if
                     os.path.isdir(os.path.join(template_path, name))]
        form_dirs.sort()
    form_index = 1
    for cur_form_dir in form_dirs:
        start_form_id = global_index
        form_name = os.path.basename(cur_form_dir)
        print("Entering " + form_name)
        field_dirs = [os.path.join(cur_form_dir, field) for field in os.listdir(cur_form_dir) if
                      os.path.isdir(os.path.join(cur_form_dir, field))]
        field_dirs = sorted(field_dirs, key=lambda a: int(os.path.basename(a)))
        field_index = 1

        for cur_field in field_dirs:
            combined_text = str()
            checkbox_text = str()
            start_field_id = global_index
            field_id = os.path.basename(cur_field)
            worksheet1.write(global_index, INDEX_NO, global_index, general_format)
            worksheet1.write(global_index, HIDDEN_COL, form_name, general_format)
            # worksheet1.write(global_index, TRUE_FALSE, 100, general_format)
            worksheet1.write(start_field_id, FIELD_ID, field_id, merge_format)

            # checkbox case
            checkbox_images = [os.path.join(cur_field, image) for image in os.listdir(cur_field) if
                               (image.endswith(".png") or image.endswith(".jpg")) and checkbox_keyword[
                                   0] in image and checkbox_keyword[1] not in image]
            checkbox_images.sort()
            if len(checkbox_images) > 0:
                checkbox_txt_files = [f for f in os.listdir(cur_field) if f.endswith(".txt") and "origin" not in f]
                # sort by converting into integer anything after "_" symbol in file name
                checkbox_txt_files = sorted(checkbox_txt_files, key=lambda a: int(a[a.find('_') + 1:-4]))
                if len(checkbox_txt_files) > 0:
                    # get ocr char and combine to form the whole text
                    try:
                        for i in range(0, len(checkbox_txt_files)):
                            tmp_file = open(cur_field + '/' + checkbox_txt_files[i], 'r')
                            # clear line break and EOF char
                            ocr_val = tmp_file.read().replace('\f', '')
                            checkbox_text += ocr_val
                            # worksheet1.write(global_index, LINE_OCR, combined_text, font_size_format)
                        #if len(checkbox_text) == 0:
                            #checkbox_text = "(no options selecteed)"

                    except Exception:
                        print("no ocr text")

                height_list = [1]
                # find biggest image to resize row height
                for img_index, image in enumerate(checkbox_images):
                    img = Image.open(image)
                    w, h = img.size
                    height_list.append(h)
                height_list.sort(reverse=True)
                worksheet1.set_row(global_index, height_list[0] * 0.75)

                for img_index, image in enumerate(checkbox_images):
                    img = Image.open(image)
                    w, h = img.size

                    if w and h:
                        x_scale = (LINE_IMG_WIDTH + 0.0) / w
                        # y_scale = (DEFAULT_ROW_HEIGHT + 0.0) / h

                    worksheet1.insert_image(
                        global_index, LINE_IMG,
                        image,
                        {
                            'x_offset': 30 + img_index * x_scale * w,
                            'y_offset': 20,
                            'x_scale': 0.8,
                            'y_scale': 0.8
                        })
                pass

            # dashline_ folders
            dashline_dirs = [dir for dir in os.listdir(cur_field)
                             if os.path.isdir(os.path.join(cur_field, dir))
                             and any(str in dir for str in dashline_ext) and mixedline_keyword not in dir]

            if len(dashline_dirs) > 0:
                for dir in dashline_dirs:
                    field_dashline_dir = os.path.join(cur_field, dir)
                    # check if dashline exist then continue
                    if os.path.exists(field_dashline_dir):
                        word_ocr_files = [f for f in os.listdir(field_dashline_dir) if f.endswith(".txt")]
                        # sort by converting into integer anything after "_" symbol in file name
                        if "dashline" in field_dashline_dir:
                            word_ocr_files = sorted(word_ocr_files, key=lambda a: int(a[a.find('_') + 1:-4]))
                        else:
                            word_ocr_files = sorted(word_ocr_files, key=lambda a: int(a[a.find('.') + 1:-4]))
                        if len(word_ocr_files) > 0:
                            # get ocr char and combine to form the whole text
                            try:
                                for i in range(0, len(word_ocr_files)):
                                    tmp_file = open(field_dashline_dir + '/' + word_ocr_files[i], 'r')
                                    # clear line break and EOF char
                                    ocr_val = tmp_file.read().replace('\n', '').replace('\f', '')
                                    combined_text += ocr_val
                            except Exception:
                                print("no ocr text")
                            # Write continuously word images to sheet
                            """for j in range(0, len(word_ocr_files)):
                                img = Image.open(field_dashline_dir + '/' + word_ocr_files[j][:-4] + ".png")
                                w, h = img.size
                                x_scale, y_scale = 0.5, 0.5

                                if w and h:
                                    x_scale = (40 + 0.0) / w
                                    y_scale = (DEFAULT_ROW_HEIGHT + 0.0) / h
                                #print(str(x_scale) + "------" + str(y_scale))
                                # write image
                                worksheet1.insert_image(
                                    global_index, LINE_IMG,
                                    field_dashline_dir + '/' + word_ocr_files[j][:-4] + ".png",
                                    {
                                        'x_offset': 0 + j * x_scale * w,
                                        'y_offset': 5,
                                        'x_scale': x_scale,
                                        'y_scale': 0.4
                                    })
                                    """
                            img = Image.open(cur_field + '/' + dir + ".png")
                            w, h = img.size
                            x_scale, y_scale = 0.5, 0.5

                            if w and h:
                                x_scale = (40 + 0.0) / w
                                y_scale = (DEFAULT_ROW_HEIGHT - 5.0) / h
                            worksheet1.insert_image(
                                global_index, LINE_IMG,
                                cur_field + '/' + dir + ".png",
                                {
                                    'x_offset': 10,
                                    'y_offset': 5,
                                    'x_scale': y_scale,
                                    'y_scale': y_scale
                                })

                            # merging field_ID cell
                            end_field_id = global_index - 1
                            if (end_field_id - start_field_id) > 0:
                                worksheet1.merge_range(start_field_id, FIELD_ID, end_field_id, FIELD_ID, field_id,
                                                       merge_format)
                            else:
                                worksheet1.write(start_field_id, FIELD_ID, field_id, merge_format)
                        else:
                            # handle empty case
                            worksheet1.write(global_index, LINE_IMG, "(empty)",
                                                 gray_format)

            # mixed line case
            mixedline_dirs = [dir for dir in os.listdir(cur_field)
                              if os.path.isdir(os.path.join(cur_field, dir))
                              and mixedline_keyword in dir]

            if len(mixedline_dirs) > 0:
                for dir in mixedline_dirs:
                    mixed_dir = os.path.join(cur_field, dir)
                    # check if mixed dir exist then continue, otherwise only print image
                    if os.path.exists(mixed_dir):
                        word_ocr_files = [f for f in os.listdir(mixed_dir) if f.endswith(".txt")]
                        # sort by converting into integer anything after "_" symbol in file name
                        word_ocr_files = sorted(word_ocr_files, key=lambda a: int(a[a.find('.') + 1:-4]))
                        if len(word_ocr_files) > 0:
                            # get ocr char and combine to form the whole text
                            try:
                                for i in range(0, len(word_ocr_files)):
                                    tmp_file = open(mixed_dir + '/' + word_ocr_files[i], 'r')
                                    # clear line break and EOF char
                                    ocr_val = tmp_file.read().replace('\n', '').replace('\f', '')
                                    combined_text += ocr_val
                            except Exception:
                                print("no ocr text")
                            # Write continuously word images to sheet
                            """for j in range(0, len(word_ocr_files)):
                                img = Image.open(mixed_dir + '/' + word_ocr_files[j][:-4] + ".png")
                                w, h = img.size
                                x_scale, y_scale = 0.5, 0.5

                                if w and h:
                                    x_scale = (40 + 0.0) / w
                                    y_scale = (DEFAULT_ROW_HEIGHT + 0.0) / h
                                #print(str(x_scale) + "------" + str(y_scale))
                                # write image
                                worksheet1.insert_image(
                                    global_index, LINE_IMG,
                                    mixed_dir + '/' + word_ocr_files[j][:-4] + ".png",
                                    {
                                        'x_offset': 0 + j * x_scale * w,
                                        'y_offset': 5,
                                        'x_scale': x_scale,
                                        'y_scale': 0.4
                                    })
                                    """
                            img = Image.open(cur_field + '/' + dir + ".png")
                            w, h = img.size
                            x_scale, y_scale = 0.5, 0.5

                            if w and h:
                                x_scale = (40 + 0.0) / w
                                y_scale = (DEFAULT_ROW_HEIGHT + 30.0) / h
                            worksheet1.insert_image(
                                global_index, LINE_IMG,
                                cur_field + '/' + dir + ".png",
                                {
                                    'x_offset': 10,
                                    'y_offset': 5,
                                    'x_scale': y_scale,
                                    'y_scale': y_scale
                                })

            worksheet1.write(global_index, LINE_OCR, combined_text + " \n " + checkbox_text, font_size_format)

            # handle empty case
            if len(os.listdir(cur_field)) == 0:
                worksheet1.write(global_index, LINE_IMG, "(empty)",
                                 gray_format)

            field_index += 1
            global_index += 1
        # merging file name cell
        end_form_id = global_index - 1
        if (end_form_id - start_form_id) > 0:
            worksheet1.merge_range(start_form_id, FILE_NAME, end_form_id, FILE_NAME, form_name, merge_format)
        else:
            worksheet1.write(start_form_id, FILE_NAME, form_name, merge_format)
        form_index += 1

    workbook.close()

    print("***********END*************")
    return


TEMPLATE01_FIELD_COUNT = 16
TEMPLATE02_FIELD_COUNT = 20
TEMPLATE03_FIELD_COUNT = 28
TEMPLATE04_FIELD_COUNT = 28
START_ROW = 2


def calculate_field_accr(xls_file):
    if len(xls_file) == 0:
        print("Empty input file")
        return

    workbook = openpyxl.load_workbook(xls_file)
    sheet_names = workbook.get_sheet_names()
    by_text_sheet = workbook.get_sheet_by_name(sheet_names[1])
    # by_text_sheet.merge_cells(start_row=2, start_column=2, end_row=2+ TEMPLATE01_FIELD_COUNT - 1, end_column=2)

    for index in range(0, TEMPLATE01_FIELD_COUNT):
        by_text_sheet.cell(row=START_ROW + index, column=2).value = "Template01"
        by_text_sheet.cell(row=START_ROW + index, column=3).value = index + 1

    for index in range(0, TEMPLATE02_FIELD_COUNT):
        by_text_sheet.cell(row=START_ROW + TEMPLATE01_FIELD_COUNT + index, column=2).value = "Template02"
        by_text_sheet.cell(row=START_ROW + TEMPLATE01_FIELD_COUNT + index, column=3).value = index + 1

    for index in range(0, TEMPLATE03_FIELD_COUNT):
        by_text_sheet.cell(row=START_ROW + TEMPLATE01_FIELD_COUNT + TEMPLATE02_FIELD_COUNT + index,
                           column=2).value = "Template03"
        by_text_sheet.cell(row=START_ROW + TEMPLATE01_FIELD_COUNT + TEMPLATE02_FIELD_COUNT + index,
                           column=3).value = index + 1

    for index in range(0, TEMPLATE04_FIELD_COUNT):
        by_text_sheet.cell(
            row=START_ROW + TEMPLATE01_FIELD_COUNT + TEMPLATE02_FIELD_COUNT + TEMPLATE03_FIELD_COUNT + index,
            column=2).value = "Template04"
        by_text_sheet.cell(
            row=START_ROW + TEMPLATE01_FIELD_COUNT + TEMPLATE02_FIELD_COUNT + TEMPLATE03_FIELD_COUNT + index,
            column=3).value = index + 1

    # 1: index #
    # 2: form name
    # 3: field ID
    # 4: field image
    # 5: ocr result
    # 6: True False
    # 7: Accuracy
    first_sheet = sheet_names[0]
    ocr_sheet = workbook.get_sheet_by_name(first_sheet)
    form_name_col = ocr_sheet.get_squared_range(HIDDEN_COL + 1, 2, HIDDEN_COL + 1, ocr_sheet.max_row)

    all_form_name_value = list()
    all_form_name_value.append("zero_indexed")
    all_form_name_value.append("title_row")
    for row in form_name_col:
        all_form_name_value.append(row[0].value)

    # find changing of template position
    changing_pos = list()
    for i in range(2, len(all_form_name_value) - 1):
        if all_form_name_value[i][0:4] != all_form_name_value[i + 1][0:4]:
            changing_pos.append(i)
    print(changing_pos)
    field_id_col = ocr_sheet.get_squared_range(FIELD_ID + 1, 2, FIELD_ID + 1, ocr_sheet.max_row)
    accuracy_col = ocr_sheet.get_squared_range(TRUE_FALSE + 1, 2, TRUE_FALSE + 1, ocr_sheet.max_row)

    all_field_ID_value = list()
    all_field_ID_value.append("zero_indexed")
    all_field_ID_value.append("title_row")
    for row in field_id_col:
        all_field_ID_value.append(row[0].value)

    all_acc_col = list()
    all_acc_col.append("zero_indexed")
    all_acc_col.append("title_row")
    for row in accuracy_col:
        all_acc_col.append(row[0].value)
    temp1_stat = [0] * (TEMPLATE01_FIELD_COUNT + 1)
    temp2_stat = [0] * (TEMPLATE02_FIELD_COUNT + 1)
    temp3_stat = [0] * (TEMPLATE03_FIELD_COUNT)
    temp4_stat = [0] * (TEMPLATE04_FIELD_COUNT)
    # for row_index in range(START_ROW, changing_pos[0]):
    offset = 1
    offset_count = 0
    for field_id in range(1, TEMPLATE01_FIELD_COUNT + 1):  # range(1...) bc Template01 has field_ID starts at 1
        # print(field_id)
        offset_count = 0
        while True:
            cur_row = offset + field_id + TEMPLATE01_FIELD_COUNT * offset_count
            if cur_row > changing_pos[0]:
                break
            if int(all_field_ID_value[cur_row]) == field_id and all_acc_col[cur_row] == 100:
                temp1_stat[field_id] += 1
            offset_count += 1
    print("OFFSET1 = " + str(offset_count))
    print(temp1_stat)

    offset = changing_pos[0]
    for field_id in range(1, TEMPLATE02_FIELD_COUNT + 1):
        offset_count = 0
        while True:
            cur_row = offset + field_id + TEMPLATE02_FIELD_COUNT * offset_count
            if cur_row > changing_pos[1]:
                break
            if int(all_field_ID_value[cur_row]) == field_id and all_acc_col[cur_row] == 100:
                temp2_stat[field_id] += 1
            offset_count += 1
    print("OFFSET2 = " + str(offset_count))
    print(temp2_stat)

    offset = changing_pos[1]
    for field_id in range(0, TEMPLATE03_FIELD_COUNT + 1):  # range(0...) bc Template03 has field_ID starts at 0
        offset_count = 0
        if field_id in [15, 16, 17]:
            continue
        while True:
            # NO FIELD ID DATA IN INPUT
            if field_id < 15:
                cur_row = offset + 1 + field_id + (TEMPLATE03_FIELD_COUNT - 3) * offset_count
            else:
                cur_row = offset + 1 + field_id - 3 + (TEMPLATE03_FIELD_COUNT - 3) * offset_count

            if cur_row > changing_pos[2]:
                break
            if int(all_field_ID_value[cur_row]) == field_id and all_acc_col[cur_row] == 100:
                temp3_stat[field_id] += 1
            offset_count += 1
    print("OFFSET3 = " + str(offset_count))
    print(temp3_stat)

    offset = changing_pos[2]
    for field_id in range(1, TEMPLATE04_FIELD_COUNT + 1):
        offset_count = 0
        # 7 12 17
        while True:
            if 12 > field_id > 7:
                cur_row = offset + field_id - 1 + TEMPLATE04_FIELD_COUNT * offset_count
            elif 17 > field_id > 12:
                cur_row = offset + field_id - 2 + TEMPLATE04_FIELD_COUNT * offset_count
            elif TEMPLATE04_FIELD_COUNT > field_id > 17:
                cur_row = offset + field_id - 3 + TEMPLATE04_FIELD_COUNT * offset_count
            else:
                cur_row = offset + field_id + TEMPLATE04_FIELD_COUNT * offset_count

            if cur_row > ocr_sheet.max_row:
                break
            if int(all_field_ID_value[cur_row]) == field_id and all_acc_col[cur_row] == 100:
                temp4_stat[field_id] += 1
            offset_count += 1
    print("OFFSET4 = " + str(offset_count))
    print(temp4_stat)

    workbook.save(xls_file[:-5] + "_acc.xlsx")
    print("-------------END-----------")
    pass


# calculate_field_accr("/home/hiepnp/Downloads/Cinnamon/Nissay/data/171004/run_4_template_2_Oct/NISSAY_FULL.xlsx")
#gen_report_file("NISSAY_TEMPLATE01", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/",
#                "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/report/Template01")

# gen_report_file("NISSAY_TEMPLATE02", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/report/Template02")

# gen_report_file("NISSAY_TEMPLATE03", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/report/Template03")

# gen_report_file("NISSAY_TEMPLATE04", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171005/report/Template04")
