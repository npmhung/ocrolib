import os
import xlsxwriter
import sys
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment

# for avoiding error while exporting Japanese
if sys.version[0] == '2':
    reload(sys)
    sys.setdefaultencoding('utf8')

# Col constant name for sheet 1
INDEX_NO = 0
FILE_NAME = 1
FIELD_ID = 2
LINE_IMG = 3
LINE_OCR = 4
TRUE_FALSE = 5
ACCURACY = 6
FIELD_TYPE = 6

# Col constant for by text sheet
TEXT_FILE_NAME = 1
TEXT_FIELD_ID = 2
TEXT_IMAGE = 3
TEXT_TRUE_FALSE = 4
TEXT_ANSWER =5
TEXT_PROBABILITY = 6

# Col constant for sheet 2
S2_FIELD_NAME = 3
S2_TOTAL = 4
S2_PREDICTED = 5
S2_ACCURACY = 6

# Default scale value
X_SCALE = 0.6
Y_SCALE = 0.5

# Rough value of converting from image dimension to xls row width(100 pixel * 0.75 will roughly fill col of 100 in xls)
PIXEL_TO_ROW_CONST = 0.8

# Field types
FIELD_NUMBER = "NUMBER"
FIELD_TEXT = "TEXT"
FIELD_CHECK_BOX = "CHECKBOX"

HIDDEN_COL = 100

DEFAULT_ROW_HEIGHT = 50
LINE_IMG_WIDTH = 90

CHECKBOX_DELIMITER = "_"
WORD_DELIMITER = "."
DASHLINE_DELIMITER = "_"
NORMAL_LINE_DELIMITER = "."

TEXT_KEYWORD = "text"
NUMBER_KEYWORD = "number"
PROBABILITY_KEYWORD = "probability"
DASHLINE_KEYWORD = ["line_", "dashline_number", "dashline_text"]
CHECKBOX_KEYWORD = ["checkbox", "origin"]
MIXED_LINE_KEYWORD = "mixedline"


def replace(parent):
    """
    :param parent:
        the path of parent dir to iterate throught
    :return:
        replace spaces ' ' and brackets '(' and ')' with underscore '_'
    """

    for path, folders, files in os.walk(parent):
        for f in files:
            os.rename(os.path.join(path, f),
                      os.path.join(path, f.replace(' ', '_').replace('(', '_').replace(')', '_')))

        for i in range(len(folders)):
            new_name = folders[i].replace(' ', '_').replace('(', '_').replace(')', '_')
            os.rename(os.path.join(path, folders[i]), os.path.join(path, new_name))
            folders[i] = new_name

def nomalize_output_text(text):
    output = []
    text = text.decode('utf-8')
    for i, c in enumerate(text):
        if ord(c) == ord(u'\u3099'):  # dakuten mark
            if i > 0: output[-1] = unichr(ord(text[i-1])+1)
        elif ord(c) == ord(u'\u309A'):  # handakuten mark
            if i > 0: output[-1] = unichr(ord(text[i-1])+2)
        else:
            output.append(c)
    return ''.join(output)


def gen_report_file(prefer_name, output_path, data_path_list):
    """
    :param prefer_name: prefer name of output .xls file
    :param output_path: prefer output dir of .xls file
    :param data_path: input of data
    :return:
    """
    # form_name = os.path.basename(data_path)

    # init workbook obj with preferred name
    workbook = xlsxwriter.Workbook(output_path + prefer_name + ".xlsx")
    # add worksheets to the workbook
    by_field_ws = workbook.add_worksheet("By Field")
    by_text_ws = workbook.add_worksheet("By Text")

    # format the font in a cell
    font_size_format = workbook.add_format()
    font_size_format.set_font_size(12)
    font_size_format.set_align('center')
    font_size_format.set_align('vcenter')
    font_size_format.set_border()
    font_size_format.set_text_wrap(True)

    general_format = workbook.add_format()
    general_format.set_align('center')
    general_format.set_align('vcenter')
    general_format.set_border()
    general_format.set_text_wrap(True)

    # Create a format to use in the merged range.
    merge_format = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })

    # set the width of columns in sheet 1
    by_field_ws.set_column(FILE_NAME, FILE_NAME, 20)
    by_field_ws.set_column(FIELD_ID, FIELD_ID, 7)
    by_field_ws.set_column(LINE_IMG, LINE_IMG, LINE_IMG_WIDTH)
    by_field_ws.set_column(LINE_OCR, LINE_OCR, 30)
    by_field_ws.set_column(TRUE_FALSE, TRUE_FALSE, 15)
    by_field_ws.set_column(FIELD_TYPE, FIELD_TYPE, 30)
    # set default height of all row of sheet 1
    by_field_ws.set_default_row(DEFAULT_ROW_HEIGHT)
    # create the title header of the sheet
    by_field_ws.write(0, INDEX_NO, "#")
    by_field_ws.write(0, FILE_NAME, "FILE NAME")
    by_field_ws.write(0, FIELD_ID, "FIELD ID")
    by_field_ws.write(0, LINE_IMG, "TEXT OR NUMBER")
    by_field_ws.write(0, LINE_OCR, "LINE OCR")
    by_field_ws.write(0, TRUE_FALSE, "TRUE OR FALSE")
    by_field_ws.write(0, FIELD_TYPE, "FIELD TYPE")

    # set the width of columns in sheet 2
    by_text_ws.set_column(TEXT_FILE_NAME, TEXT_FILE_NAME, 30)
    by_text_ws.set_column(TEXT_FIELD_ID, TEXT_FIELD_ID, 10)
    by_text_ws.set_column(TEXT_IMAGE, TEXT_IMAGE, 40)
    by_text_ws.set_column(TEXT_TRUE_FALSE, TEXT_TRUE_FALSE, 30)
    by_text_ws.set_column(TEXT_ANSWER, TEXT_ANSWER, 20)
    by_text_ws.set_column(TEXT_PROBABILITY, TEXT_PROBABILITY, 20)
    # set default height of all row of sheet 2
    by_text_ws.set_default_row(50)
    # create the title header of the sheet
    by_text_ws.write(0, TEXT_FILE_NAME, "FILE NAME")
    by_text_ws.write(0, TEXT_FIELD_ID, "FIELD ID")
    by_text_ws.write(0, TEXT_IMAGE, "TEXT OR NUMBER (SINGLE)")
    by_text_ws.write(0, TEXT_TRUE_FALSE, "TRUE OR FALSE")
    by_text_ws.write(0, TEXT_ANSWER, "ANSWER")
    by_text_ws.write(0, TEXT_PROBABILITY, "PROBABILITY")

    # fix spaces in paths
    for data_path in data_path_list: replace(data_path)
    form_dirs = data_path_list

    template_index = 1
    global_index = 1
    form_index = 1
    by_text_sheet_index = 1

    for cur_form_dir in form_dirs:
        form_name = os.path.basename(cur_form_dir)
        field_dirs = [os.path.join(cur_form_dir, field) for field in os.listdir(cur_form_dir) if
                      os.path.isdir(os.path.join(cur_form_dir, field))]
        field_dirs = sorted(field_dirs, key=lambda a: int(os.path.basename(a)))
        field_index = 1

        for cur_field in field_dirs:
            # reset row size and offset values
            x_offset = 5
            y_offset = 5
            # list to keep track of current images height and width
            height_list = [1]
            width_list = [1]

            field_type_text = set()
            combined_text = str()
            checkbox_text = str()
            start_field_id = global_index
            field_id = os.path.basename(cur_field)
            by_field_ws.write(global_index, INDEX_NO, global_index, general_format)
            by_field_ws.write(global_index, HIDDEN_COL, form_name, general_format)
            by_field_ws.write(start_field_id, FIELD_ID, field_id, merge_format)

            # checkbox case
            height_list[:] = []
            width_list[:] = []
            checkbox_images = [os.path.join(cur_field, image) for image in os.listdir(cur_field) if
                               (image.endswith(".png") or image.endswith(".jpg")) and CHECKBOX_KEYWORD[0]
                               in image and CHECKBOX_KEYWORD[1] not in image]
            checkbox_images.sort()
            if len(checkbox_images) > 0:
                field_type_text.add(FIELD_CHECK_BOX)
                checkbox_txt_files = [f for f in os.listdir(cur_field)
                                      if f.endswith(".txt") and CHECKBOX_KEYWORD[1] not in f
                                      and PROBABILITY_KEYWORD not in f]
                # sort by converting into integer anything after "_" symbol in file name
                checkbox_txt_files = sorted(checkbox_txt_files,
                                            key=lambda a: int(a[a.find(CHECKBOX_DELIMITER) + 1:-4]))
                if len(checkbox_txt_files) > 0:
                    # get ocr char and combine to form the whole text
                    try:
                        for i in range(0, len(checkbox_txt_files)):
                            tmp_file = open(cur_field + '/' + checkbox_txt_files[i], 'r')
                            # clear line break and EOF char
                            ocr_val = tmp_file.read().replace('\f', '')
                            checkbox_text += ocr_val
                            # by_field_ws.write(global_index, LINE_OCR, combined_text, font_size_format)
                    except Exception:
                        print("No ocr text")

                # find biggest image to resize row height
                for img_index, image in enumerate(checkbox_images):
                    img = Image.open(image)
                    w, h = img.size
                    height_list.append(h)
                    width_list.append(w)
                height_list.sort(reverse=True)
                width_list.sort(reverse=True)

                # more than 2 and big checkbox images (eg: template04, field24)
                # check width and insert vertically
                if len(checkbox_images) > 2 and width_list[0] / 2 > LINE_IMG_WIDTH:
                    for img_index, image in enumerate(checkbox_images):
                        img = Image.open(image)
                        w, h = img.size
                        by_field_ws.insert_image(
                            global_index, LINE_IMG,
                            image,
                            {
                                'x_offset': x_offset,
                                'y_offset': y_offset,
                                'x_scale': 0.8,
                                'y_scale': Y_SCALE
                            })
                        # x_offset not changed
                        y_offset += Y_SCALE * h
                else:  # normal cases, insert horizontally
                    for img_index, image in enumerate(checkbox_images):
                        img = Image.open(image)
                        w, h = img.size
                        height_list.append(h)
                        if w / 2 > DEFAULT_ROW_HEIGHT:
                            x_scale = 0.7
                        else:
                            x_scale = 1
                        by_field_ws.insert_image(
                            global_index, LINE_IMG,
                            image,
                            {
                                'x_offset': x_offset,
                                'y_offset': y_offset,
                                'x_scale': x_scale,
                                'y_scale': 1
                            })
                        x_offset += x_scale * w
                    # y_offset changed
                    height_list.sort(reverse=True)
                    if len(height_list):
                        y_offset += height_list[0]
                pass

            # dashline_folders
            height_list[:] = []
            width_list[:] = []
            dashline_dirs = [dir for dir in os.listdir(cur_field)
                             if os.path.isdir(os.path.join(cur_field, dir))
                             and any(str in dir for str in DASHLINE_KEYWORD) and MIXED_LINE_KEYWORD not in dir]
            if len(dashline_dirs) > 0:
                for dashline_index, dir in enumerate(dashline_dirs):
                    x_offset = 5
                    field_dashline_dir = os.path.join(cur_field, dir)
                    # check if dashline exist then continue
                    if os.path.exists(field_dashline_dir):
                        word_ocr_files = [f for f in os.listdir(field_dashline_dir)
                                          if f.endswith(".txt") and PROBABILITY_KEYWORD not in f]
                        if TEXT_KEYWORD.lower() in field_dashline_dir.lower():
                            field_type_text.add(FIELD_TEXT)
                        if NUMBER_KEYWORD.lower() in field_dashline_dir.lower():
                            field_type_text.add(FIELD_NUMBER)
                        # sort by converting into integer anything after "_" symbol in file name
                        if "dashline" in field_dashline_dir:
                            word_ocr_files = sorted(word_ocr_files,
                                                    key=lambda a: int(a[a.find(DASHLINE_DELIMITER) + 1:-4]))
                        else:
                            word_ocr_files = sorted(word_ocr_files,
                                                    key=lambda a: int(a[a.find(NORMAL_LINE_DELIMITER) + 1:-4]))
                        if len(word_ocr_files) > 0:
                            # get ocr char and combine to form the whole text
                            try:
                                for i in range(0, len(word_ocr_files)):
                                    tmp_file = open(field_dashline_dir + '/' + word_ocr_files[i], 'r')
                                    # clear line break and EOF char
                                    ocr_val = tmp_file.read().replace('\n', '').replace('\f', '')

                                    ############################################
                                    # by text sheet, insert letter image and text
                                    insert_by_text_row(by_text_ws, by_text_sheet_index, form_name, field_id,
                                                       field_dashline_dir, word_ocr_files[i], font_size_format)
                                    by_text_sheet_index += 1
                                    #############################################
                                    combined_text += ocr_val
                                # not the last index, add line break
                                if dashline_index != len(dashline_dirs) - 1:
                                    combined_text += "\n"
                            except Exception:
                                print("No ocr text")

                            # Write continuously word images to by field sheet
                            for j in range(0, len(word_ocr_files)):
                                img = Image.open(field_dashline_dir + '/' + word_ocr_files[j][:-4] + ".png")
                                w, h = img.size
                                height_list.append(h)
                                x_scale, y_scale = 0.5, 0.5

                                if w and h:
                                    x_scale = (40 + 0.0) / w
                                    y_scale = (40 + 0.0) / h
                                # write image
                                by_field_ws.insert_image(
                                    global_index, LINE_IMG,
                                    field_dashline_dir + '/' + word_ocr_files[j][:-4] + ".png",
                                    {
                                        'x_offset': x_offset,
                                        'y_offset': y_offset,
                                        'x_scale': x_scale,
                                        'y_scale': y_scale
                                    })
                                x_offset += x_scale * w


                    height_list.sort(reverse=True)
                    if len(height_list):
                        y_offset += height_list[0]

            # mixed line case
            height_list[:] = []
            width_list[:] = []
            mixedline_dirs = [dir for dir in os.listdir(cur_field)
                              if os.path.isdir(os.path.join(cur_field, dir))
                              and MIXED_LINE_KEYWORD in dir]
            if len(mixedline_dirs) > 0:
                field_type_text.add(FIELD_NUMBER)
                for mixedline_index, dir in enumerate(mixedline_dirs):
                    x_offset = 5
                    mixed_dir = os.path.join(cur_field, dir)
                    # check if mixed dir exist then continue, otherwise only print image
                    if os.path.exists(mixed_dir):
                        word_ocr_files = [f for f in os.listdir(mixed_dir) if f.endswith(".txt")
                                          and PROBABILITY_KEYWORD not in f]
                        # sort by converting into integer anything after "_" symbol in file name
                        word_ocr_files = sorted(word_ocr_files, key=lambda a: int(a[a.find(WORD_DELIMITER) + 1:-4]))
                        if len(word_ocr_files) > 0:
                            # get ocr char and combine to form the whole text
                            try:
                                for i in range(0, len(word_ocr_files)):
                                    tmp_file = open(mixed_dir + '/' + word_ocr_files[i], 'r')
                                    # clear line break and EOF char
                                    ocr_val = tmp_file.read().replace('\n', '').replace('\f', '')

                                    ############################################
                                    # by text sheet, insert letter image and text
                                    insert_by_text_row(by_text_ws, by_text_sheet_index, form_name, field_id,
                                                       mixed_dir, word_ocr_files[i], font_size_format)
                                    by_text_sheet_index += 1
                                    #############################################

                                    combined_text += ocr_val
                                # not the last index, add line break
                                if mixedline_index != len(mixedline_dirs) - 1:
                                    combined_text += "\n"
                            except Exception:
                                print("no ocr text")
                            # Write continuously word images to sheet
                            for j in range(0, len(word_ocr_files)):
                                img = Image.open(mixed_dir + '/' + word_ocr_files[j][:-4] + ".png")
                                w, h = img.size
                                height_list.append(h)
                                if w and h:
                                    x_scale = (40 + 0.0) / w
                                # write image
                                by_field_ws.insert_image(
                                    global_index, LINE_IMG,
                                    mixed_dir + '/' + word_ocr_files[j][:-4] + ".png",
                                    {
                                        'x_offset': x_offset,
                                        'y_offset': y_offset,
                                        'x_scale': x_scale,
                                        'y_scale': Y_SCALE
                                    })
                                x_offset += x_scale * w
                    height_list.sort(reverse=True)
                    if len(height_list):
                        y_offset += height_list[0]
            by_field_ws.write(global_index, LINE_OCR, checkbox_text + " \n " + nomalize_output_text(combined_text) , font_size_format)
            by_field_ws.write(global_index, FILE_NAME, form_name, font_size_format)
            temp_str = ''
            for index, item in enumerate(field_type_text):
                temp_str += item + "\n"
            by_field_ws.write(global_index, FIELD_TYPE, temp_str, font_size_format)
            if y_offset > DEFAULT_ROW_HEIGHT:
                by_field_ws.set_row(global_index, y_offset)
            else:
                by_field_ws.set_row(global_index, DEFAULT_ROW_HEIGHT)
            field_index += 1
            global_index += 1

        form_index += 1
        template_index += 1
    workbook.close()
    print("***********SUCCESS*************")
    return

def gen_report_file_original(prefer_name, output_path, data_path_list):
    """
    :param prefer_name: prefer name of output .xls file
    :param output_path: prefer output dir of .xls file
    :param data_path: input of data
    :return:
    """
    # form_name = os.path.basename(data_path)

    # init workbook obj with preferred name
    workbook = xlsxwriter.Workbook(output_path + prefer_name + ".xlsx")
    # add worksheets to the workbook
    by_field_ws = workbook.add_worksheet("By Field")
    by_text_ws = workbook.add_worksheet("By Text")

    # format the font in a cell
    font_size_format = workbook.add_format()
    font_size_format.set_font_size(12)
    font_size_format.set_align('center')
    font_size_format.set_align('vcenter')
    font_size_format.set_border()
    font_size_format.set_text_wrap(True)

    general_format = workbook.add_format()
    general_format.set_align('center')
    general_format.set_align('vcenter')
    general_format.set_border()
    general_format.set_text_wrap(True)

    # Create a format to use in the merged range.
    merge_format = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })

    # set the width of columns in sheet 1
    by_field_ws.set_column(FILE_NAME, FILE_NAME, 20)
    by_field_ws.set_column(FIELD_ID, FIELD_ID, 7)
    by_field_ws.set_column(LINE_IMG, LINE_IMG, LINE_IMG_WIDTH)
    by_field_ws.set_column(LINE_OCR, LINE_OCR, 30)
    by_field_ws.set_column(TRUE_FALSE, TRUE_FALSE, 15)
    by_field_ws.set_column(FIELD_TYPE, FIELD_TYPE, 30)
    # set default height of all row of sheet 1
    by_field_ws.set_default_row(DEFAULT_ROW_HEIGHT)
    # create the title header of the sheet
    by_field_ws.write(0, INDEX_NO, "#")
    by_field_ws.write(0, FILE_NAME, "FILE NAME")
    by_field_ws.write(0, FIELD_ID, "FIELD ID")
    by_field_ws.write(0, LINE_IMG, "TEXT OR NUMBER")
    by_field_ws.write(0, LINE_OCR, "LINE OCR")
    by_field_ws.write(0, TRUE_FALSE, "TRUE OR FALSE")
    by_field_ws.write(0, FIELD_TYPE, "FIELD TYPE")

    # set the width of columns in sheet 2
    by_text_ws.set_column(TEXT_FILE_NAME, TEXT_FILE_NAME, 30)
    by_text_ws.set_column(TEXT_FIELD_ID, TEXT_FIELD_ID, 10)
    by_text_ws.set_column(TEXT_IMAGE, TEXT_IMAGE, 40)
    by_text_ws.set_column(TEXT_TRUE_FALSE, TEXT_TRUE_FALSE, 30)
    by_text_ws.set_column(TEXT_ANSWER, TEXT_ANSWER, 20)
    by_text_ws.set_column(TEXT_PROBABILITY, TEXT_PROBABILITY, 20)
    # set default height of all row of sheet 2
    by_text_ws.set_default_row(50)
    # create the title header of the sheet
    by_text_ws.write(0, TEXT_FILE_NAME, "FILE NAME")
    by_text_ws.write(0, TEXT_FIELD_ID, "FIELD ID")
    by_text_ws.write(0, TEXT_IMAGE, "TEXT OR NUMBER (SINGLE)")
    by_text_ws.write(0, TEXT_TRUE_FALSE, "TRUE OR FALSE")
    by_text_ws.write(0, TEXT_ANSWER, "ANSWER")
    by_text_ws.write(0, TEXT_PROBABILITY, "PROBABILITY")

    # fix spaces in paths
    for data_path in data_path_list: replace(data_path)
    form_dirs = data_path_list

    template_index = 1
    global_index = 1
    form_index = 1
    by_text_sheet_index = 1

    for cur_form_dir in form_dirs:
        form_name = os.path.basename(cur_form_dir)
        field_dirs = [os.path.join(cur_form_dir, field) for field in os.listdir(cur_form_dir) if
                      os.path.isdir(os.path.join(cur_form_dir, field))]
        field_dirs = sorted(field_dirs, key=lambda a: int(os.path.basename(a)))
        field_index = 1

        for cur_field in field_dirs:
            # reset row size and offset values
            x_offset = 5
            y_offset = 5
            # list to keep track of current images height and width
            height_list = [1]
            width_list = [1]

            field_type_text = set()
            combined_text = str()
            checkbox_text = str()
            start_field_id = global_index
            field_id = os.path.basename(cur_field)
            by_field_ws.write(global_index, INDEX_NO, global_index, general_format)
            by_field_ws.write(global_index, HIDDEN_COL, form_name, general_format)
            by_field_ws.write(start_field_id, FIELD_ID, field_id, merge_format)

            # checkbox case
            height_list[:] = []
            width_list[:] = []
            checkbox_images = [os.path.join(cur_field, image) for image in os.listdir(cur_field) if
                               (image.endswith(".png") or image.endswith(".jpg")) and CHECKBOX_KEYWORD[0]
                               in image and CHECKBOX_KEYWORD[1] not in image]
            checkbox_images.sort()
            if len(checkbox_images) > 0:
                field_type_text.add(FIELD_CHECK_BOX)
                checkbox_txt_files = [f for f in os.listdir(cur_field)
                                      if f.endswith(".txt") and CHECKBOX_KEYWORD[1] not in f
                                      and PROBABILITY_KEYWORD not in f]
                # sort by converting into integer anything after "_" symbol in file name
                checkbox_txt_files = sorted(checkbox_txt_files,
                                            key=lambda a: int(a[a.find(CHECKBOX_DELIMITER) + 1:-4]))
                if len(checkbox_txt_files) > 0:
                    # get ocr char and combine to form the whole text
                    try:
                        for i in range(0, len(checkbox_txt_files)):
                            tmp_file = open(cur_field + '/' + checkbox_txt_files[i], 'r')
                            # clear line break and EOF char
                            ocr_val = tmp_file.read().replace('\f', '')
                            checkbox_text += ocr_val
                            # by_field_ws.write(global_index, LINE_OCR, combined_text, font_size_format)
                    except Exception:
                        print("No ocr text")

                # find biggest image to resize row height
                for img_index, image in enumerate(checkbox_images):
                    img = Image.open(image)
                    w, h = img.size
                    height_list.append(h)
                    width_list.append(w)
                height_list.sort(reverse=True)
                width_list.sort(reverse=True)

                # more than 2 and big checkbox images (eg: template04, field24)
                # check width and insert vertically
                if len(checkbox_images) > 2 and width_list[0] / 2 > LINE_IMG_WIDTH:
                    for img_index, image in enumerate(checkbox_images):
                        img = Image.open(image)
                        w, h = img.size
                        by_field_ws.insert_image(
                            global_index, LINE_IMG,
                            image,
                            {
                                'x_offset': x_offset,
                                'y_offset': y_offset,
                                'x_scale': 0.8,
                                'y_scale': Y_SCALE
                            })
                        # x_offset not changed
                        y_offset += Y_SCALE * h
                else:  # normal cases, insert horizontally
                    for img_index, image in enumerate(checkbox_images):
                        img = Image.open(image)
                        w, h = img.size
                        height_list.append(h)
                        if w / 2 > DEFAULT_ROW_HEIGHT:
                            x_scale = 0.7
                        else:
                            x_scale = 1
                        by_field_ws.insert_image(
                            global_index, LINE_IMG,
                            image,
                            {
                                'x_offset': x_offset,
                                'y_offset': y_offset,
                                'x_scale': x_scale,
                                'y_scale': 1
                            })
                        x_offset += x_scale * w
                    # y_offset changed
                    height_list.sort(reverse=True)
                    if len(height_list):
                        y_offset += height_list[0]
                pass

            # dashline_folders
            height_list[:] = []
            width_list[:] = []
            dashline_dirs = [dir for dir in os.listdir(cur_field)
                             if os.path.isdir(os.path.join(cur_field, dir))
                             and any(str in dir for str in DASHLINE_KEYWORD) and MIXED_LINE_KEYWORD not in dir]
            if len(dashline_dirs) > 0:
                for dashline_index, dir in enumerate(dashline_dirs):
                    x_offset = 5
                    field_dashline_dir = os.path.join(cur_field, dir)
                    # check if dashline exist then continue
                    if os.path.exists(field_dashline_dir):
                        word_ocr_files = [f for f in os.listdir(field_dashline_dir)
                                          if f.endswith(".txt") and PROBABILITY_KEYWORD not in f]
                        if TEXT_KEYWORD.lower() in field_dashline_dir.lower():
                            field_type_text.add(FIELD_TEXT)
                        if NUMBER_KEYWORD.lower() in field_dashline_dir.lower():
                            field_type_text.add(FIELD_NUMBER)
                        # sort by converting into integer anything after "_" symbol in file name
                        if "dashline" in field_dashline_dir:
                            word_ocr_files = sorted(word_ocr_files,
                                                    key=lambda a: int(a[a.find(DASHLINE_DELIMITER) + 1:-4]))
                        else:
                            word_ocr_files = sorted(word_ocr_files,
                                                    key=lambda a: int(a[a.find(NORMAL_LINE_DELIMITER) + 1:-4]))
                        if len(word_ocr_files) > 0:
                            # get ocr char and combine to form the whole text
                            try:
                                for i in range(0, len(word_ocr_files)):
                                    tmp_file = open(field_dashline_dir + '/' + word_ocr_files[i], 'r')
                                    # clear line break and EOF char
                                    ocr_val = tmp_file.read().replace('\n', '').replace('\f', '')

                                    ############################################
                                    # by text sheet, insert letter image and text
                                    insert_by_text_row(by_text_ws, by_text_sheet_index, form_name, field_id,
                                                       field_dashline_dir, word_ocr_files[i], font_size_format)
                                    by_text_sheet_index += 1
                                    #############################################
                                    combined_text += ocr_val
                                # not the last index, add line break
                                if dashline_index != len(dashline_dirs) - 1:
                                    combined_text += "\n"
                            except Exception:
                                print("No ocr text")

                            # Write continuously word images to by field sheet
                            for j in range(0, len(word_ocr_files)):
                                img = Image.open(field_dashline_dir + '/' + word_ocr_files[j][:-4] + ".png")
                                w, h = img.size
                                height_list.append(h)
                                x_scale, y_scale = 0.5, 0.5

                                if w and h:
                                    x_scale = (40 + 0.0) / w
                                    y_scale = (40 + 0.0) / h
                                # write image
                                by_field_ws.insert_image(
                                    global_index, TEXT_TRUE_FALSE,
                                    field_dashline_dir + '/' + word_ocr_files[j][:-4] + ".png",
                                    {
                                        'x_offset': x_offset,
                                        'y_offset': y_offset,
                                        'x_scale': x_scale,
                                        'y_scale': y_scale
                                    })
                                x_offset += x_scale * w

                    # insert dashline image
                    img = Image.open(field_dashline_dir + ".png")
                    w, h = img.size
                    height_list.append(h)
                    x_offset, y_offset = 0.5, 0.5
                    by_field_ws.insert_image(global_index, LINE_IMG, field_dashline_dir + '.png',
                                             {
                                                 'x_offset': x_offset,
                                                 'y_offset': y_offset,
                                                 'x_scale': 0.7,
                                                 'y_scale': 0.7
                                             })
                    height_list.sort(reverse=True)
                    if len(height_list):
                        y_offset += height_list[0]

            # mixed line case
            height_list[:] = []
            width_list[:] = []
            mixedline_dirs = [dir for dir in os.listdir(cur_field)
                              if os.path.isdir(os.path.join(cur_field, dir))
                              and MIXED_LINE_KEYWORD in dir]
            if len(mixedline_dirs) > 0:
                field_type_text.add(FIELD_NUMBER)
                for mixedline_index, dir in enumerate(mixedline_dirs):
                    x_offset = 5
                    mixed_dir = os.path.join(cur_field, dir)
                    # check if mixed dir exist then continue, otherwise only print image
                    if os.path.exists(mixed_dir):
                        word_ocr_files = [f for f in os.listdir(mixed_dir) if f.endswith(".txt")
                                          and PROBABILITY_KEYWORD not in f]
                        # sort by converting into integer anything after "_" symbol in file name
                        word_ocr_files = sorted(word_ocr_files, key=lambda a: int(a[a.find(WORD_DELIMITER) + 1:-4]))
                        if len(word_ocr_files) > 0:
                            # get ocr char and combine to form the whole text
                            try:
                                for i in range(0, len(word_ocr_files)):
                                    tmp_file = open(mixed_dir + '/' + word_ocr_files[i], 'r')
                                    # clear line break and EOF char
                                    ocr_val = tmp_file.read().replace('\n', '').replace('\f', '')

                                    ############################################
                                    # by text sheet, insert letter image and text
                                    insert_by_text_row(by_text_ws, by_text_sheet_index, form_name, field_id,
                                                       mixed_dir, word_ocr_files[i], font_size_format)
                                    by_text_sheet_index += 1
                                    #############################################

                                    combined_text += ocr_val
                                # not the last index, add line break
                                if mixedline_index != len(mixedline_dirs) - 1:
                                    combined_text += "\n"
                            except Exception:
                                print("no ocr text")

                            # Write continuously word images to sheet
                            # for j in range(0, len(word_ocr_files)):
                            #     img = Image.open(mixed_dir + '/' + word_ocr_files[j][:-4] + ".png")
                            #     w, h = img.size
                            #     height_list.append(h)
                            #     if w and h:
                            #         x_scale = (40 + 0.0) / w
                            #     # write image
                            #     by_field_ws.insert_image(
                            #         global_index, LINE_IMG,
                            #         mixed_dir + '/' + word_ocr_files[j][:-4] + ".png",
                            #         {
                            #             'x_offset': x_offset,
                            #             'y_offset': y_offset,
                            #             'x_scale': x_scale,
                            #             'y_scale': Y_SCALE
                            #         })
                            #     x_offset += x_scale * w

                    # insert dashline image
                    img = Image.open(mixed_dir + ".png")
                    w, h = img.size
                    height_list.append(h)
                    by_field_ws.insert_image(global_index, LINE_IMG, mixed_dir + '.png',
                                             {
                                                 'x_offset': x_offset,
                                                 'y_offset': y_offset,
                                                 'x_scale': 0.7,
                                                 'y_scale': 0.7
                                             })
                    height_list.sort(reverse=True)
                    if len(height_list):
                        y_offset += height_list[0]
            by_field_ws.write(global_index, LINE_OCR, checkbox_text + " \n " + nomalize_output_text(combined_text) , font_size_format)
            by_field_ws.write(global_index, FILE_NAME, form_name, font_size_format)
            temp_str = ''
            for index, item in enumerate(field_type_text):
                temp_str += item + "\n"
            by_field_ws.write(global_index, FIELD_TYPE, temp_str, font_size_format)
            if y_offset > DEFAULT_ROW_HEIGHT:
                by_field_ws.set_row(global_index, y_offset)
            else:
                by_field_ws.set_row(global_index, DEFAULT_ROW_HEIGHT)
            field_index += 1
            global_index += 1

        form_index += 1
        template_index += 1
    workbook.close()
    print("***********SUCCESS*************")
    return

START_ROW = 2


def insert_by_text_row(worksheet, row_index, file_name, field_id, dir, txt_file, format):
    '''
    Function to insert into by text sheet
    :param worksheet: worksheet to insert
    :param row_index: index of row
    :param file_name: current form name to print (dansin0001 ...)
    :param field_id: field id for print
    :param dir: directory contains input files (source image, orc text, probability text files)
    :param txt_file: we use text ocr file as source of
    :param format:
    :return:
    '''
    with open(dir + '/' + txt_file, 'r') as f:
        # clear line break and EOF char
        ocr_val = f.read().replace('\n', '').replace('\f', '')

    image = dir + '/' + txt_file[:-4] + ".png"
    img = Image.open(image)
    w, h = img.size
    if w and h:
        x_scale = (40 + 0.0) / w
        y_scale = (40 + 0.0) / h
    prob_txt = dir + '/' + txt_file[:-4] + "_" + PROBABILITY_KEYWORD + ".txt"
    try:
        with open(prob_txt) as f:
            prob_val = f.read().replace('\n', '').replace('\f', '')
    except Exception:
        prob_val = "0"
    worksheet.write(row_index, TEXT_FILE_NAME, file_name, format)
    worksheet.write(row_index, TEXT_FIELD_ID, field_id, format)
    worksheet.insert_image(row_index, TEXT_IMAGE, image, {
                                        'x_offset': 5,
                                        'y_offset': 5,
                                        'x_scale': x_scale,
                                        'y_scale': y_scale
                                    })
    worksheet.write(row_index, TEXT_ANSWER, ocr_val, format)
    worksheet.write(row_index, TEXT_PROBABILITY, prob_val)
    pass


def file_len(fname):
    with open(fname) as f:
        for i, l in enumerate(f):
            pass
    return i + 1


def calculate_field_accr(data_path, xls_file_name, template_name):
    print("Calculating accuracy of template report...")
    if len(xls_file_name) == 0:
        print("Empty input file")
        return
    # read field names from txt file
    field_names = []
    try:
        with open(os.path.join(data_path + "/" + template_name + ".txt")) as f:
            data = f.read()
            field_names = data.split("\n")
    except Exception:
        print("No field names file found, please specify field names file to calculate result")
        return
    field_count = file_len(os.path.join(data_path + "/" + template_name + ".txt"))
    try:
        workbook = openpyxl.load_workbook(os.path.join(data_path, xls_file_name) + ".xlsx")
    except Exception:
        print("Cant find report xls file to calculate data")
        return

    sheet_names = workbook.get_sheet_names()
    by_text_sheet = workbook.get_sheet_by_name(sheet_names[1])

    try:
        for index in range(0, field_count):
            by_text_sheet.cell(row=START_ROW + index, column=S2_FIELD_NAME).value = field_names[index]
    except IndexError:
        print(IndexError)
    # 1: index #
    # 2: form name
    # 3: field ID
    # 4: field image
    # 5: ocr result
    # 6: True False
    # 7: Accuracy
    first_sheet = sheet_names[0]
    ocr_sheet = workbook.get_sheet_by_name(first_sheet)
    form_name_col = ocr_sheet.get_squared_range(HIDDEN_COL + 1, 2, HIDDEN_COL + 1, ocr_sheet.max_row)

    all_form_name_value = list()
    all_form_name_value.append("zero_indexed")
    all_form_name_value.append("title_row")
    for row in form_name_col:
        all_form_name_value.append(row[0].value)

    field_id_col = ocr_sheet.get_squared_range(FIELD_ID + 1, 2, FIELD_ID + 1, ocr_sheet.max_row)
    predict_result_col = ocr_sheet.get_squared_range(TRUE_FALSE + 1, 2, TRUE_FALSE + 1, ocr_sheet.max_row)

    all_field_id_value = list()
    all_field_id_value.append("zero_indexed")
    all_field_id_value.append("title_row")
    for row in field_id_col:
        all_field_id_value.append(row[0].value)

    all_acc_col = list()
    all_acc_col.append("zero_indexed")
    all_acc_col.append("title_row")
    for row in predict_result_col:
        all_acc_col.append(row[0].value)

    temp_stat = [0] * (field_count + 1)
    temp_all_occurences = [0] * (field_count + 1)
    field_id_list = [number for number in range(0, field_count + 1)]

    for row in range(2, ocr_sheet.max_row + 1):
        if len(str(all_acc_col[row])) > 0:
            if int(all_field_id_value[row]) in field_id_list:
                temp_all_occurences[int(all_field_id_value[row])] += 1
                if str(all_acc_col[row]).upper() == "T":
                    temp_stat[int(all_field_id_value[row])] += 1

    for index in range(0, field_count):
        by_text_sheet.cell(row=START_ROW + index, column=S2_TOTAL).value = str(temp_all_occurences[index])
        by_text_sheet.cell(row=START_ROW + index, column=S2_PREDICTED).value = str(temp_stat[index])
        acc_val = 0
        if temp_all_occurences[index] != 0:
            acc_val = 100 * (int(temp_stat[index]) + 0.0) / int(temp_all_occurences[index])
        by_text_sheet.cell(row=START_ROW + index, column=S2_ACCURACY).value = str(acc_val)
    workbook.remove(ocr_sheet)
    workbook.save(data_path + "/" + xls_file_name + "_FIELD_ACCR.xlsx")
    print("-------------SUCCESS-----------")
    return


# data_path = ["/home/hiepnp/Downloads/Cinnamon/Nissay/data/171010/Template03/sougou0001"]
# gen_report_file("sougou0001", "/home/hiepnp/Downloads/Cinnamon/Nissay/data/171010/", data_path)
# calculate_field_accr("/home/hiepnp/Downloads/Cinnamon/Nissay/data/171011/nissay_report_171011_filled", "Template01", "Template01")
# calculate_field_accr("/home/hiepnp/Downloads/Cinnamon/Nissay/data/171011/nissay_report_171011_filled", "Template02", "Template02")
# calculate_field_accr("/home/hiepnp/Downloads/Cinnamon/Nissay/data/171011/nissay_report_171011_filled", "Template03", "Template03")
# calculate_field_accr("/home/hiepnp/Downloads/Cinnamon/Nissay/data/171011/nissay_report_171011_filled", "Template04", "Template04")

